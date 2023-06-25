import openpyxl
import datetime
import os

def Read_Book(in_file):
    dirname = os.path.dirname(in_file)
    cbook = openpyxl.load_workbook(in_file)
    print(in_file)
    sheet = cbook['Sheet1']
    i = 1
    f_count = 1
    dt_now = datetime.datetime.now()
    while True:
        if sheet.cell(row = i, column = 1).value == None:
            break
        elif sheet.cell(row = i, column = 2).value == None:
            f_name = str(dt_now.year) + f'{dt_now.month:02}' + f'{dt_now.day:02}' + '_' + f'{dt_now.hour:02}' + f'{dt_now.minute:02}' + f'{dt_now.second:02}' + '_' +  f'{f_count:04}' + '.txt'
            f = open(dirname + '/' + f_name, 'w', encoding='utf-8')
            f.writelines(sheet.cell(row = i, column = 1).value)
            f.close()
            sheet.cell(row = i, column = 2).value = True
            f_count += 1
        i += 1
    cbook.save(in_file)